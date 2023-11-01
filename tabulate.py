from openpyxl import Workbook
from openpyxl.styles import PatternFill

row_colors = {
  "meh": [
    PatternFill(start_color=color, fill_type="solid") for color in [
      "ffbcb8",
      "ffa099",
      "ff7066",
      "ff3c2e"
    ]
  ],
  "decent": [
    PatternFill(start_color=color, fill_type="solid") for color in [
      "ffe4b8",
      "ffd28a",
      "ffbe57",
      "ffab24"
    ]
  ],
  "good": [
    PatternFill(start_color=color, fill_type="solid") for color in [
      "a0ff9e",
      "d2ffd1",
      "04FA00",
      "03CC00"
    ]
  ]
}

def num_to_col(num):
  col_str = chr(ord('A') + (num % 26))
  num //= 26
  
  while num > 26:
    col_str = chr(ord('A') + (num % 26)) + col_str
    num //= 26
  if num > 0:
    col_str = chr(ord('A') + (num % 26 - 1)) + col_str
  
  return col_str

def fill_cells(cells, color):
   for row in cells:
        for cell in row:
          cell.fill = color


def generate_spreadsheet(filename, document_metadata):
  wb = Workbook()
  ws = wb.active
  row = 1
  for key in document_metadata.keys():
    ws[f'A{row}'] = key
    ws.merge_cells(f'A{row}:D{row}')
    ws[f'A{row}'].fill = row_colors[key][3]
    
    row += 1
    for i, val in enumerate(['Title', 'URL', 'OneDrive Link', 'Doing']):
      ws[f'{num_to_col(i)}{row}'] = val
    fill_cells(ws[f'A{row}:D{row}'], row_colors[key][2])  
      
    row += 1
    for item in document_metadata[key]:
      for i, val in enumerate([item.title, item.url]):
        ws[f'{num_to_col(i)}{row}'] = val
      fill_cells(ws[f'A{row}:D{row}'], row_colors[key][row % 2])
      row += 1
    row += 1
    
    wb.save(filename)
    